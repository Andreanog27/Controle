from django.shortcuts import render, redirect, get_object_or_404
import json
import csv
import calendar
from datetime import datetime
from django.db.models import Sum
from django.contrib.auth.forms import UserCreationForm
from .forms import RegistroForm
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.urls import path, include
from .models import Despesa, Categoria
from .forms import DespesaForm
from .models import Receita
from .forms import ReceitaForm
from django.contrib import messages
from django.db.models.functions import TruncMonth
from decimal import Decimal
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db.models.functions import ExtractMonth
from .models import Despesa, Categoria
from django.shortcuts import redirect
from django.utils.timezone import now
from datetime import timedelta

@login_required
def index(request):
    despesas = Despesa.objects.filter(usuario=request.user).order_by('-data')
    categorias = Categoria.objects.all()

    mes = request.GET.get('mes')
    ano = request.GET.get('ano')
    categoria = request.GET.get('categoria')

    if mes:
        despesas = despesas.filter(data__month=mes)
    if ano:
        despesas = despesas.filter(data__year=ano)
    if categoria:
        despesas = despesas.filter(categoria_id=categoria)

    total = despesas.aggregate(Sum("valor"))["valor__sum"] or 0

    return render(request, "index.html", {
        "despesas": despesas,
        "total": total,
        "categorias": categorias,
    })

@login_required
def adicionar_despesa(request):
    print('cheguei aqui: 1')
    if request.method == "POST":
        print('cheguei aqui: 2')
        print(request.POST)
        form = DespesaForm(request.POST)
        print("Erros do form:", form.errors)

        if form.is_valid():
            print('cheguei aqui: 3')
            despesa = form.save(commit=False)
            despesa.usuario = request.user  # IMPORTANTE!
            despesa.save()
            return redirect('index')
    else:
        form = DespesaForm()

    return render(request, 'adicionar_despesa.html', {'form': form})

@login_required
def listar_despesas(request):
    despesas = Despesa.objects.filter(usuario=request.user).order_by('-data')
    return render(request, 'despesas/listar_despesas.html', {"despesas": despesas})

def nova_despesa(request):
    categorias = Categoria.objects.all()

    if request.method == 'POST':
        descricao = request.POST['descricao']
        valor = request.POST['valor']
        categoria_id = request.POST['categoria']  

        categoria = get_object_or_404(Categoria, id=categoria_id)
        

        Despesa.objects.create(
            descricao=descricao,
            valor=valor,
            categoria=categoria,
            data=datetime.today(),
            usuario=request.user
        )

        return redirect('dashboard')

    return render(request, 'nova_despesa.html', {'categorias': categorias})

def editar_despesa(request, id):
    despesa = get_object_or_404(Despesa, id=id)

    if request.method == "POST":
        despesa.descricao = request.POST.get("descricao")
        despesa.valor = request.POST.get("valor")
        despesa.data = request.POST.get("data")
        despesa.categoria_id = request.POST.get("categoria")
        despesa.save()
        return redirect("dashboard")

    categorias = Categoria.objects.all()
    return render(request, "despesas/editar_despesa.html", {
        "despesa": despesa,
        "categorias": categorias
    })

@login_required
def adicionar_receita(request):
    if request.method == "POST":
        form = ReceitaForm(request.POST)
        if form.is_valid():
            receita = form.save(commit=False)
            receita.usuario = request.user
            receita.save()
            return redirect('index')
    else:
        form = ReceitaForm()

    return render(request, "adicionar_receita.html", {"form": form})

@login_required
def listar_receitas(request):
    receitas = Receita.objects.filter(usuario=request.user).order_by('-data')
    return render(request, 'receitas/listar_receitas.html', {"receitas": receitas})


def editar_receita(request, id):
    receita = get_object_or_404(Receita, id=id)

    if request.method == "POST":
        receita.descricao = request.POST.get("descricao")
        receita.valor = request.POST.get("valor")
        receita.data = request.POST.get("data")
        receita.save()
        return redirect("listar_receitas")

    return render(request, "receitas/editar_receita.html", {
        "receita": receita
    })

@login_required
def excluir_receita(request, id):
    # Busca a receita do usuário logado
    receita = get_object_or_404(Receita, id=id)
    
    if request.method == "POST":
        receita.delete()
        messages.success(request, "Receita excluída com sucesso.")
    
    # Sempre redireciona para o dashboard
    return redirect('dashboard')

def relatorios(request):
    categorias = Categoria.objects.all()
    despesas = Despesa.objects.filter(usuario=request.user)

    # Pegando filtros da URL (GET)
    categoria_id = request.GET.get("categoria")
    data_inicio = request.GET.get("data_inicio")
    data_fim = request.GET.get("data_fim")

    # Aplicando filtros
    if categoria_id and categoria_id.isdigit():
        despesas = despesas.filter(categoria_id=categoria_id)
    
    if data_inicio:
        despesas = despesas.filter(data__gte=data_inicio)
    
    if data_fim:
        despesas = despesas.filter(data__lte=data_fim)

    return render(request, "relatorios.html", {
        "categorias": categorias,
        "despesas": despesas,
        "filtro_categoria": int(categoria_id) if categoria_id else None,
        "filtro_data_inicio": data_inicio,
        "filtro_data_fim": data_fim,
    })
     

def fazer_login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user:
            login(request, user)
            return redirect('index')
        else:
            messages.error(request, "Usuário ou senha inválidos.")

    return render(request, 'login.html')


def registrar(request):
    if request.method == "POST":
        form = RegistroForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data["password"])
            user.save()
            messages.success(request, "Conta criada com sucesso! Faça login.")
            return redirect('login')
    else:
        form = RegistroForm()

    return render(request, 'registrar.html')


def sair(request):
    logout(request)
    return redirect('login')

@login_required
def dashboard(request):
    usuario = request.user

    despesas = Despesa.objects.filter(usuario=usuario).order_by("data")
    receitas = Receita.objects.filter(usuario=usuario).order_by("data")

    total_gastos = despesas.aggregate(total=Sum("valor"))["total"] or 0
    total_receitas = receitas.aggregate(total=Sum("valor"))["total"] or 0

    saldo = total_receitas - total_gastos

    # --- CÁLCULO DA MÉDIA MENSAL ---
    if despesas.exists():
        primeiro_registro = despesas.order_by("data").first().data
        hoje = datetime.today().date()  # corrigido

        meses = (hoje.year - primeiro_registro.year) * 12 + (hoje.month - primeiro_registro.month)
        meses = max(meses, 1)  # para evitar divisão por zero

        media_mensal = total_gastos / meses
    else:
        media_mensal = 0

    # --- RECEITA MENSAL DO MÊS ATUAL ---
    hoje = datetime.today().date()  # corrigido
    receita_mensal = receitas.filter(
        data__year=hoje.year,
        data__month=hoje.month
    ).aggregate(Sum("valor"))["valor__sum"] or 0

    # --- Gráfico mensal de DESPESAS ---
    meses_labels = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
    valores_gastos = [
        float(despesas.filter(data__month=m).aggregate(total=Sum("valor"))["total"] or 0)
        for m in range(1, 13)
    ]

# --- Gráfico mensal de RECEITAS ---
    valores_receitas = [
    float(receitas.filter(data__month=m).aggregate(total=Sum("valor"))["total"] or 0)
    for m in range(1, 13)
    ]


    # --- Gráfico de pizza por categoria ---
    categorias_labels = list(Categoria.objects.values_list("nome", flat=True))
    categorias_valores = [
        float(despesas.filter(categoria__nome=nome).aggregate(total=Sum("valor"))["total"] or 0)
        for nome in categorias_labels
    ]

    context = {
        "despesas": despesas,
        "receitas": receitas,
        "total_gastos": total_gastos,
        "total_receitas": total_receitas,
        "saldo": saldo,
        "media_mensal": media_mensal,         # adicionada
        "receita_mensal": receita_mensal,     # adicionada
        "meses": meses_labels,
        "valores_gastos": valores_gastos,
        "valores_receitas": valores_receitas,   # <- adicionada
        "categorias_labels": categorias_labels,
        "categorias_valores": categorias_valores,
    }

    return render(request, "dashboard.html", context)


@login_required
def excluir_despesa(request, despesa_id):
    """
    Exclui uma despesa do usuário logado de forma segura.
    - Se a despesa não existir ou não pertencer ao usuário, mostra mensagem de erro.
    - Se excluída com sucesso, mostra mensagem de sucesso.
    """
    if request.method == 'POST':
        try:
            despesa = Despesa.objects.get(id=despesa_id, usuario=request.user)
        except Despesa.DoesNotExist:
            messages.error(request, "Despesa não encontrada ou não pertence a você.")
            return redirect('dashboard')

        despesa.delete()
        messages.success(request, f"Despesa '{despesa.descricao}' excluída com sucesso!")
        return redirect('dashboard')

    # Se acessar via GET, apenas redireciona
    messages.warning(request, "A exclusão deve ser feita via botão de confirmação.")
    return redirect('dashboard')

def excluir_categoria(request, id):
    categoria = get_object_or_404(Categoria, id=id)
    categoria.delete()
    messages.success(request, "Categoria excluída com sucesso!")
    return redirect('index')

@login_required
def exportar_excel(request):
    despesas = Despesa.objects.filter(usuario=request.user).order_by("-data")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Despesas"

    # Cabeçalho
    colunas = ["Descrição", "Valor", "Categoria", "Data"]
    ws.append(colunas)

    # Dados
    for despesa in despesas:
        ws.append([
            despesa.descricao,
            float(despesa.valor),
            despesa.categoria.nome if despesa.categoria else "",
            despesa.data.strftime("%d/%m/%Y")
        ])

    # Ajusta largura das colunas
    for col in ws.columns:
        max_len = 0
        col_letra = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[col_letra].width = max_len + 2

    # Resposta HTTP com o arquivo
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=despesas.xlsx"
    wb.save(response)

    return response

def exportar_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="despesas.csv"'

    writer = csv.writer(response)
    writer.writerow(["Descrição", "Categoria", "Valor", "Data"])

    despesas = Despesa.objects.all()
    for d in despesas:
        writer.writerow([d.descricao, d.categoria, d.valor, d.data])

    return response

